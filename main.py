import win32com.client
import openai
import pandas as pd
import re
import win32com.client
import openpyxl

email_address = "XYZ@XYZ.com" # Replace with your email address
openai.api_key = "XYZ" # Replace with your openai api key
database_link = r"XYZ.xlsx"


# Optional - read text on which chatGPT will make personalized email (for example a description of a company)
list_1 = pd.read_excel(database_link)
filtered_data = list_1[list_1['Founded'] == 2022]

# Function for sending an email
def send_outlook_email(subject, body, to_email, attachment_path):
  outlook_app = win32com.client.Dispatch("Outlook.Application")
  namespace = outlook_app.GetNamespace("MAPI")

  # Specify the account for sending the email
  accounts = namespace.Accounts
  for account in accounts:
    if account.SmtpAddress == email_address:
      sender_account = account
      break

  else:
    raise Exception("Email account not found")

  # Create a new mail item
  mail_item = outlook_app.CreateItem(0)  # 0 represents olMailItem (Mail item)

  # Set the email properties
  mail_item.Subject = "Maximize Your Finances: Explore Projections, Valuations, and ESG Reporting Support"
  mail_item.Body = body
  mail_item.To = to_email

  # Set the account for sending the email
  mail_item._oleobj_.Invoke(*(64209, 0, 8, 0, sender_account))

  # Attach the file
  mail_item.Attachments.Add(attachment_path)  # Add the attachment using the attachment_path parameter

  # Send the email
  mail_item.Send()

# Define a regular expression pattern to match the Subject line
subject_pattern = re.compile(r"Subject: (.+?)\n", re.DOTALL)

# Iterate through database with information (get out email and check if the emails is already sent)
for index, row in filtered_data.iterrows():
  email = row['Email'] # data will probably containt your email (if not, adjust to your needs)
  sent = row['Sent']
  index_nr = row['Nr'] + 1

  if pd.notnull(email) and pd.isnull(sent):
    business_desc = row['Business Description']
    startup = row['Startup']
    print(startup)

    try:
      client = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
          {"role": "system", "content": "You are marketing expert, skilled in constructing personalized email content."}, # Adjust to your needs
          {"role": "user", "content": f"Write your personalized prompt toward chatGPT" # Adjust to your needs
           }
        ]
      )

      # Define a regular expression pattern to match the Subject line
      subject_pattern = re.compile(r"Subject: (.+?)\n", re.DOTALL)

      # Search for the subject in the email text
      subject_match = subject_pattern.search(client.choices[0].message["content"])

      body = client.choices[0].message["content"]

      # If you also prompt chatGPT to make subject, if not adjust the code
      if subject_match:
        # Extract the subject
        subject = subject_match.group(1)
        print("Subject:", subject)

        # Remove the subject from the email text
        subject_match = subject_pattern.sub("", client.choices[0].message["content"])
        print("\nEmail Text without Subject:\n", subject_match)
        body = subject_match

      else:
        print("No subject found in the email.")

      # Example usage
      subject = "Maximize Your Finances: Explore Projections, Valuations, and ESG Reporting Support"

      recipient = email

      # Optional - attach something in the email
      attachment_path = r"XXX.pdf" # attachment of something useful for recipient

      send_outlook_email(subject, body, recipient, attachment_path)

      # Optional & simple line of code to keep track in database of emails sent
      wb = openpyxl.load_workbook(database_link)
      sheet = wb.active
      sheet.cell(row=index_nr, column=13, value="Sent")
      wb.save(database_link)

    except openai.error.Timeout as e:
      print(f"Timeout occurred: {e}")
      # Optional & simple line of code to keep track in database of emails sent
      wb = openpyxl.load_workbook(database_link)
      sheet = wb.active
      sheet.cell(row=index_nr, column=13, value="Timeout occurred")
      wb.save(database_link)
